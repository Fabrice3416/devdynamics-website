# Generation : python3 build_petite_caisse.py petite_caisse.xlsx ../assets/images/logo.png   (openpyxl + pillow)
"""Gabarit Excel : Fiche de suivi - Petite caisse (Bousol / KesKle)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import sys, os

OUT = sys.argv[1]
LOGO = sys.argv[2]

INK, GREY, OLIVE = "2A2A28", "6B6A66", "4C5A47"
STONE, INPUT, LINE = "EFEDE8", "FAF8F3", "C9C4BA"
BODY = "Candara"; TITLE = "Palatino Linotype"

f_body   = Font(name=BODY, size=10, color=INK)
f_small  = Font(name=BODY, size=8.5, color=GREY)
f_bold   = Font(name=BODY, size=10, bold=True, color=INK)
f_label  = Font(name=BODY, size=9.5, bold=True, color=INK)
f_org    = Font(name=TITLE, size=12, bold=True, color=OLIVE)
f_title  = Font(name=TITLE, size=16, bold=True, color=OLIVE)
f_sect   = Font(name=TITLE, size=11.5, bold=True, color=OLIVE)
f_ital   = Font(name=BODY, size=10, italic=True, color=GREY)
f_head   = Font(name=BODY, size=9, bold=True, color=INK)

fill_stone = PatternFill("solid", fgColor=STONE)
fill_input = PatternFill("solid", fgColor=INPUT)
thin  = Side(style="thin", color=LINE)
hair  = Side(style="hair", color=LINE)
med   = Side(style="medium", color=OLIVE)
box   = Border(left=thin, right=thin, top=thin, bottom=thin)
box_h = Border(left=thin, right=thin, top=thin, bottom=med)
rule  = Border(bottom=med)
NUM = '#,##0.00;-#,##0.00;;@'
DATE = 'DD/MM/YYYY'
L = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=0)
C = Alignment(horizontal="center", vertical="center", wrap_text=True)
R = Alignment(horizontal="right", vertical="center")
TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

wb = Workbook()
ws = wb.active
ws.title = "Petite caisse"
ws.sheet_view.showGridLines = False

widths = {"A": 11.5, "B": 11.5, "C": 34.5, "D": 9.5, "E": 13, "F": 13, "G": 14, "H": 10, "I": 26}
for c, w in widths.items():
    ws.column_dimensions[c].width = w

def put(ref, value=None, font=f_body, fill=None, border=None, align=L, fmt=None):
    cell = ws[ref]
    if value is not None:
        cell.value = value
    cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border
    cell.alignment = align
    if fmt: cell.number_format = fmt
    return cell

def merge(rng, **kw):
    ws.merge_cells(rng)
    first = rng.split(":")[0]
    put(first, **kw)
    # bordures sur toute la plage fusionnee
    if kw.get("border") or kw.get("fill"):
        from openpyxl.utils.cell import range_boundaries
        c1, r1, c2, r2 = range_boundaries(rng)
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = ws.cell(row=r, column=c)
                if kw.get("border"): cell.border = kw["border"]
                if kw.get("fill"): cell.fill = kw["fill"]
    return ws[first]

# ---------------- En-tete DevDynamics ----------------
# Logo sur A:B (rangees 1-2), texte a partir de C, mentions d'edition a droite.
logo = XLImage(LOGO)
ratio = logo.height / logo.width
logo.width = 162; logo.height = int(162 * ratio)
ws.add_image(logo, "A1")
merge("C1:F1", value="DÉVELOPPEMENT ET DYNAMISME", font=f_org)
put("G1", value="Exemplaire :", font=f_small, align=R)
merge("H1:I1", fill=fill_input, border=box, font=f_small, align=C)
merge("C2:I2", value="Association éducative et technologique à but non lucratif · Cap-Haïtien, Haïti · dev-dynamics.org", font=f_small)
merge("A3:F3", value="Projet KèsKlè — Programme d'Appui aux Initiatives Émergentes de la Société Civile (PAIESC), financé par l'Union européenne", font=f_small)
put("G3", value="N° de fiche :", font=f_small, align=R)
merge("H3:I3", fill=fill_input, border=box, font=f_small, align=C)
merge("A4:B4", value="Contrat de subvention n° :", font=f_body)
merge("C4:F4", fill=fill_input, border=box, font=f_body)
put("G4", value="Édité le :", font=f_small, align=R)
merge("H4:I4", fill=fill_input, border=box, font=f_small, align=C, fmt=DATE)
ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 15
ws.row_dimensions[3].height = 20
ws.row_dimensions[4].height = 20
for col in "ABCDEFGHI":
    ws[f"{col}5"].border = rule
ws.row_dimensions[5].height = 4

merge("A6:I6", value="FICHE DE SUIVI — PETITE CAISSE", font=f_title, align=C)
merge("A7:I7", value="Journal de caisse et arrêté · Fonds fixe", font=f_small, align=C)
ws.row_dimensions[6].height = 26

# ---------------- Identification ----------------
rows = [
    ("Nom de l'organisation", "DÉVELOPPEMENT ET DYNAMISME", "Période (du … au …)", None),
    ("Intitulé du projet", "KèsKlè", "Devise", "HTG (gourde)"),
    ("Responsable (RAF)", None, "Détenteur de la caisse", None),
    ("Montant initial de la caisse (fonds fixe)", None, "Plafond par dépense en espèces", None),
]
r = 9
for lab1, v1, lab2, v2 in rows:
    merge(f"A{r}:B{r}", value=lab1, font=f_label, fill=fill_stone, border=box)
    merge(f"C{r}:D{r}", value=v1, font=f_body, fill=(None if v1 else fill_input), border=box)
    merge(f"E{r}:F{r}", value=lab2, font=f_label, fill=fill_stone, border=box)
    merge(f"G{r}:I{r}", value=v2, font=f_body, fill=(None if v2 else fill_input), border=box)
    ws.row_dimensions[r].height = 27
    r += 1
ws["C12"].number_format = NUM; ws["C12"].alignment = R
ws["G12"].number_format = NUM; ws["G12"].alignment = R

# ---------------- Journal ----------------
H = 14
heads = ["Date", "N° pièce", "Intitulé de la dépense", "Ligne budg.", "Entrée (HTG)", "Sortie (HTG)", "Balance (HTG)", "Dép. reportée", "Observations"]
for i, h in enumerate(heads, start=1):
    put(f"{get_column_letter(i)}{H}", value=h, font=f_head, fill=fill_stone, border=box_h, align=C)
ws.row_dimensions[H].height = 30

FIRST, LAST = 16, 45
# Solde initial
put(f"A{H+1}", border=box); put(f"B{H+1}", border=box)
put(f"C{H+1}", value="Solde initial (report ou fonds fixe)", font=f_ital, border=box)
put(f"D{H+1}", border=box); put(f"E{H+1}", border=box); put(f"F{H+1}", border=box)
put(f"G{H+1}", value="=C12", font=f_bold, fill=fill_input, border=box, align=R, fmt=NUM)
put(f"H{H+1}", border=box); put(f"I{H+1}", border=box)

for rr in range(FIRST, LAST + 1):
    put(f"A{rr}", fill=fill_input, border=box, align=C, fmt=DATE)
    put(f"B{rr}", fill=fill_input, border=box, align=C)
    put(f"C{rr}", fill=fill_input, border=box)
    put(f"D{rr}", fill=fill_input, border=box, align=C)
    put(f"E{rr}", fill=fill_input, border=box, align=R, fmt=NUM)
    put(f"F{rr}", fill=fill_input, border=box, align=R, fmt=NUM)
    put(f"G{rr}", value=f'=IF(AND(E{rr}="",F{rr}=""),"",$G$15+SUM($E$16:E{rr})-SUM($F$16:F{rr}))',
        border=box, align=R, fmt=NUM)
    put(f"H{rr}", fill=fill_input, border=box, align=C)
    put(f"I{rr}", fill=fill_input, border=box)
    ws.row_dimensions[rr].height = 17

dv = DataValidation(type="list", formula1='"Oui,Non"', allow_blank=True, showDropDown=False)
dv.error = "Saisir Oui ou Non"; dv.errorTitle = "Dépense reportée"
ws.add_data_validation(dv); dv.add(f"H{FIRST}:H{LAST}")
dvn = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
dvn.error = "Montant positif en gourdes"; dvn.errorTitle = "Montant"
ws.add_data_validation(dvn); dvn.add(f"E{FIRST}:F{LAST}")

T = LAST + 1
merge(f"A{T}:D{T}", value="TOTAUX DE LA PÉRIODE", font=f_bold, fill=fill_stone, border=box)
put(f"E{T}", value=f"=SUM(E{FIRST}:E{LAST})", font=f_bold, fill=fill_stone, border=box, align=R, fmt=NUM)
put(f"F{T}", value=f"=SUM(F{FIRST}:F{LAST})", font=f_bold, fill=fill_stone, border=box, align=R, fmt=NUM)
put(f"G{T}", value=f"=G15+E{T}-F{T}", font=f_bold, fill=fill_stone, border=box, align=R, fmt=NUM)
merge(f"H{T}:I{T}", fill=fill_stone, border=box)
ws.row_dimensions[T].height = 20

r = T + 2  # 48
merge(f"A{r}:C{r}", value="Solde de caisse à la date du :", font=f_label)
put(f"D{r}", fill=fill_input, border=box, align=C, fmt=DATE)
merge(f"E{r}:F{r}", value="Solde de caisse (HTG) :", font=f_label, align=R)
put(f"G{r}", value=f"=G{T}", font=f_bold, border=box, align=R, fmt=NUM)
r += 1
merge(f"A{r}:I{r}", value=("Toute dépense en espèces est appuyée d'un justificatif et d'un reçu ; aucune dépense ne dépasse le plafond unitaire ; "
                          "le renflouement n'intervient qu'après justification des dépenses antérieures et arrêté de caisse daté et signé. "
                          "Les chèques de renflouement sont émis au nom du détenteur nommément désigné, jamais au porteur."),
      font=f_small, align=TOP)
ws.row_dimensions[r].height = 34

# ---------------- Arrete ----------------
r += 2  # 51
merge(f"A{r}:I{r}", value="ARRÊTÉ DE CAISSE", font=f_sect, border=rule)
ws.row_dimensions[r].height = 22
A = r + 1
arr = [
    ("Date de l'arrêté", None, "Solde théorique (journal)", f"=G{T}"),
    ("Espèces comptées (solde constaté)", None, "Écart (constaté − théorique)", f"=IF(C{A+1}=\"\",\"\",C{A+1}-G{A})"),
    ("Montant du renflouement demandé", f"=IF(C{A+1}=\"\",\"\",MAX(0,C12-C{A+1}))", "Commentaire (obligatoire si écart)", None),
]
for i, (lab1, v1, lab2, v2) in enumerate(arr):
    rr = A + i
    merge(f"A{rr}:B{rr}", value=lab1, font=f_label, fill=fill_stone, border=box)
    c = merge(f"C{rr}:D{rr}", value=v1, font=f_body, fill=(fill_input if v1 is None else None), border=box, align=R)
    merge(f"E{rr}:F{rr}", value=lab2, font=f_label, fill=fill_stone, border=box)
    g = merge(f"G{rr}:I{rr}", value=v2, font=f_body, fill=(fill_input if v2 is None else None), border=box, align=(L if i == 2 else R))
    ws.row_dimensions[rr].height = 27
ws[f"C{A}"].number_format = DATE; ws[f"C{A}"].alignment = C
for ref in (f"C{A+1}", f"C{A+2}", f"G{A}", f"G{A+1}"):
    ws[ref].number_format = NUM

# ---------------- Signatures ----------------
r = A + 4  # 56
merge(f"A{r}:C{r}", value="Détenteur de la caisse", font=f_label, fill=fill_stone, border=box, align=C)
merge(f"D{r}:F{r}", value="Établi par le Responsable Administratif et Financier", font=f_label, fill=fill_stone, border=box, align=C)
merge(f"G{r}:I{r}", value="Visa du Coordinateur", font=f_label, fill=fill_stone, border=box, align=C)
ws.row_dimensions[r].height = 26
merge(f"A{r+1}:C{r+1}", value="Nom :", font=f_small, border=box, align=TOP)
merge(f"D{r+1}:F{r+1}", value="Nom :", font=f_small, border=box, align=TOP)
merge(f"G{r+1}:I{r+1}", value="Nom :", font=f_small, border=box, align=TOP)
ws.row_dimensions[r+1].height = 18
merge(f"A{r+2}:C{r+2}", value="Date et signature :", font=f_small, border=box, align=TOP)
merge(f"D{r+2}:F{r+2}", value="Date et signature :", font=f_small, border=box, align=TOP)
merge(f"G{r+2}:I{r+2}", value="Date et signature :", font=f_small, border=box, align=TOP)
ws.row_dimensions[r+2].height = 58

# ---------------- Legende / pied ----------------
r += 4  # 60
put(f"A{r}", fill=fill_input, border=box)
merge(f"B{r}:D{r}", value="Cellule à saisir", font=f_small)
put(f"E{r}", border=box)
merge(f"F{r}:I{r}", value="Formule automatique — ne pas modifier", font=f_small)
merge(f"A{r+1}:I{r+1}", value="Bousòl · Fiche conforme à la feuille « Suivi Petite Caisse » du formulaire de suivi des dépenses du PAIESC, complétée du n° de pièce, de la ligne budgétaire et de l'arrêté de caisse (guide de procédures, § 4.5 du cahier des charges).",
      font=f_small, align=TOP)
ws.row_dimensions[r+1].height = 26

# ---------------- Impression ----------------
ws.print_area = f"A1:I{r+1}"
ws.page_setup.orientation = "portrait"
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
ws.print_options.horizontalCentered = True
ws.oddFooter.center.text = "Petite caisse — &A — page &P/&N"
ws.oddFooter.center.size = 8
ws.freeze_panes = "A15"

# ---------------- Notice ----------------
n = wb.create_sheet("Notice")
n.sheet_view.showGridLines = False
n.column_dimensions["A"].width = 30
n.column_dimensions["B"].width = 90
n["A1"].value = "Notice d'utilisation"; n["A1"].font = f_sect
items = [
    ("Objet", "Journal de la petite caisse en fonds fixe, tenu par le RAF, arrêté à chaque fin de période et avant tout renflouement."),
    ("Solde initial", "Au premier mois : le fonds fixe (case C12). Ensuite : le solde constaté à l'arrêté précédent, à saisir en G15 (remplace la formule)."),
    ("N° pièce", "Numéro attribué par Bousòl au règlement : rubrique-séquence (ex. 03-014 = 14e pièce de la rubrique Bureau local)."),
    ("Ligne budg.", "Une seule ligne budgétaire par dépense (ex. 3.1, 3.2, 5.2). Une dépense couvrant deux lignes est scindée en deux."),
    ("Entrée", "Uniquement les renflouements (chèque au nom du détenteur désigné) ou une régularisation d'écart."),
    ("Sortie", "Dépense en espèces appuyée d'un justificatif et d'un reçu ; jamais au-dessus du plafond unitaire (case G12)."),
    ("Balance", "Calculée automatiquement : solde initial + entrées − sorties."),
    ("Dép. reportée", "Oui si le justificatif de la dépense n'est pas encore versé à la date du mouvement ; Non sinon."),
    ("Arrêté de caisse", "Compter les espèces, saisir le solde constaté ; l'écart et le renflouement demandé (fonds fixe − constaté) se calculent. Un écart impose un commentaire."),
    ("Signatures", "Détenteur, RAF (établit), Coordinateur (visa). En régime papier : trois exemplaires, un pour l'organisation, deux pour le bailleur."),
]
for i, (k, v) in enumerate(items, start=3):
    n[f"A{i}"].value = k; n[f"A{i}"].font = f_label; n[f"A{i}"].alignment = TOP
    n[f"B{i}"].value = v; n[f"B{i}"].font = f_body; n[f"B{i}"].alignment = TOP
    n.row_dimensions[i].height = 30
n.page_setup.orientation = "landscape"; n.page_setup.paperSize = n.PAPERSIZE_A4
n.page_setup.fitToWidth = 1; n.sheet_properties.pageSetUpPr.fitToPage = True

wb.properties.title = "Fiche de suivi - Petite caisse"
wb.properties.creator = "Bousòl - DÉVELOPPEMENT ET DYNAMISME"
wb.save(OUT)
print("OK", OUT)
